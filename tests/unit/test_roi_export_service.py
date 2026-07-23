"""CSV/Excel output, manifest, overwrite, and partial-failure tests."""

from __future__ import annotations

import csv
import io
import warnings
from pathlib import Path

import pytest
from openpyxl import load_workbook

from iclotspython.output.contracts import (
    ExportFormat,
    ExportRequest,
    FailurePolicy,
    ManifestStatus,
    OutputFileStatus,
    OverwritePolicy,
)
from iclotspython.output.errors import (
    InvalidDestinationError,
    InvalidOutputRequestError,
)
from iclotspython.output.exporting import export_roi_data
from tests.support.output_factory import (
    roi_result,
    workspace_output_directory,
)


pytestmark = pytest.mark.unit


def test_csv_export_content_manifest_and_requested_files_only():
    with workspace_output_directory() as destination:
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                formats=(ExportFormat.CSV,),
            )
        )
        assert manifest.status is ManifestStatus.SUCCESS
        record = manifest.files[0]
        assert record.format == "csv"
        assert record.status is OutputFileStatus.CREATED
        path = Path(record.path)
        with path.open(newline="", encoding="utf-8") as handle:
            rows = list(csv.DictReader(handle))
        assert rows[0]["frame_label"] == "baseline"
        assert rows[2]["red_area_um2"] == "0.5"
        assert {item.name for item in destination.iterdir()} == {path.name}


def test_excel_export_has_stable_sheets_and_no_csv_side_effect():
    with workspace_output_directory() as destination:
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                formats=(ExportFormat.EXCEL,),
            )
        )
        assert manifest.status is ManifestStatus.SUCCESS
        path = Path(manifest.files[0].path)
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore", category=DeprecationWarning, module=r"openpyxl\..*"
            )
            workbook = load_workbook(
                io.BytesIO(path.read_bytes()), read_only=True, data_only=True
            )
        try:
            assert workbook.sheetnames == [
                "Data",
                "Parameters",
                "Provenance",
                "Warnings",
            ]
            data = workbook["Data"]
            assert data["A2"].value == "baseline"
            assert data["D4"].value == pytest.approx(0.5)
        finally:
            workbook.close()
        assert {item.name for item in destination.iterdir()} == {path.name}


def test_csv_and_excel_request_creates_exactly_two_manifested_files():
    with workspace_output_directory() as destination:
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                formats=(ExportFormat.CSV, ExportFormat.EXCEL),
            )
        )
        assert manifest.status is ManifestStatus.SUCCESS
        assert [record.format for record in manifest.files] == ["csv", "xlsx"]
        assert {item.name for item in destination.iterdir()} == {
            Path(record.path).name for record in manifest.generated_files
        }


def test_export_does_not_silently_overwrite():
    with workspace_output_directory() as destination:
        target = destination / "sample_ROI_roi_data.csv"
        target.write_text("original", encoding="utf-8")
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                formats=(ExportFormat.CSV,),
            )
        )
        assert manifest.status is ManifestStatus.FAILED
        assert manifest.files[0].issues[0].code == "target_exists"
        assert target.read_text(encoding="utf-8") == "original"


def test_explicit_export_replace_is_recorded():
    with workspace_output_directory() as destination:
        target = destination / "sample_ROI_roi_data.csv"
        target.write_text("original", encoding="utf-8")
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                formats=(ExportFormat.CSV,),
                overwrite=OverwritePolicy.REPLACE,
            )
        )
        assert manifest.status is ManifestStatus.SUCCESS
        assert manifest.files[0].status is OutputFileStatus.REPLACED
        assert manifest.files[0].issues[0].code == "target_replaced"
        assert target.read_text(encoding="utf-8").startswith("frame_label,time")


def test_continue_policy_returns_partial_manifest_for_one_collision():
    with workspace_output_directory() as destination:
        existing = destination / "sample_ROI_roi_data.xlsx"
        existing.write_bytes(b"original workbook")
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                formats=(ExportFormat.CSV, ExportFormat.EXCEL),
                failure_policy=FailurePolicy.CONTINUE,
            )
        )
        assert manifest.status is ManifestStatus.PARTIAL
        assert [record.status for record in manifest.files] == [
            OutputFileStatus.CREATED,
            OutputFileStatus.FAILED,
        ]
        assert manifest.files[1].issues[0].code == "target_exists"
        assert existing.read_bytes() == b"original workbook"
        assert {item.name for item in destination.iterdir()} == {
            "sample_ROI_roi_data.csv",
            "sample_ROI_roi_data.xlsx",
        }


def test_stop_policy_marks_later_formats_not_attempted():
    with workspace_output_directory() as destination:
        existing = destination / "sample_ROI_roi_data.csv"
        existing.write_text("original", encoding="utf-8")
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                formats=(ExportFormat.CSV, ExportFormat.EXCEL),
                failure_policy=FailurePolicy.STOP,
            )
        )
        assert manifest.status is ManifestStatus.FAILED
        assert [record.status for record in manifest.files] == [
            OutputFileStatus.FAILED,
            OutputFileStatus.NOT_ATTEMPTED,
        ]
        assert not (destination / "sample_ROI_roi_data.xlsx").exists()


def test_writer_failure_is_partial_and_cleans_temporary_file(monkeypatch):
    from iclotspython.output import exporting

    with workspace_output_directory() as destination:
        def fail_excel(*args, **kwargs):
            raise RuntimeError("synthetic Excel failure")

        monkeypatch.setattr(exporting, "_write_excel", fail_excel)
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                formats=(ExportFormat.CSV, ExportFormat.EXCEL),
                failure_policy=FailurePolicy.CONTINUE,
            )
        )
        assert manifest.status is ManifestStatus.PARTIAL
        assert manifest.files[1].issues[0].code == "export_write_failed"
        assert {item.name for item in destination.iterdir()} == {
            "sample_ROI_roi_data.csv"
        }


def test_invalid_destination_and_empty_or_duplicate_formats_are_structured():
    with workspace_output_directory() as parent:
        missing = parent / "missing"
        with pytest.raises(InvalidDestinationError):
            export_roi_data(
                ExportRequest(
                    result=roi_result(), destination_directory=missing
                )
            )
        with pytest.raises(InvalidOutputRequestError):
            export_roi_data(
                ExportRequest(
                    result=roi_result(),
                    destination_directory=parent,
                    formats=(),
                )
            )
        with pytest.raises(InvalidOutputRequestError):
            export_roi_data(
                ExportRequest(
                    result=roi_result(),
                    destination_directory=parent,
                    formats=(ExportFormat.CSV, ExportFormat.CSV),
                )
            )


def test_export_can_create_destination_explicitly_and_sanitize_custom_stem():
    with workspace_output_directory() as parent:
        destination = parent / "created"
        manifest = export_roi_data(
            ExportRequest(
                result=roi_result(),
                destination_directory=destination,
                create_destination=True,
                stem="custom result",
            )
        )
        assert manifest.status is ManifestStatus.SUCCESS
        assert Path(manifest.files[0].path).name == "custom_result_roi_data.csv"
