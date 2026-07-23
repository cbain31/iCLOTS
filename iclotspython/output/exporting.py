"""CSV and Excel export services for ROI accumulation results."""

from __future__ import annotations

import csv
import os
import tempfile
import warnings
from pathlib import Path

from iclotspython.presentation.tables import roi_table

from .contracts import (
    ExportFormat,
    ExportRequest,
    FailurePolicy,
    IssueSeverity,
    OutputFileRecord,
    OutputFileStatus,
    OutputIssue,
    OutputManifest,
    OverwritePolicy,
)
from .errors import InvalidOutputRequestError
from .paths import (
    build_manifest,
    prepare_destination,
    sanitize_stem,
    success_record,
)


def suggest_export_filename(result, format: ExportFormat) -> str:
    """Return a deterministic default export filename."""
    stem = sanitize_stem(result.suggested_export_stem, result.workflow_id)
    return f"{stem}_roi_data.{format.value}"


def _write_csv(result, handle) -> None:
    table = roi_table(result)
    writer = csv.writer(handle, lineterminator="\n")
    writer.writerow(table.columns)
    writer.writerows(table.rows)


def _excel_value(value):
    if isinstance(value, (tuple, list, dict)):
        return repr(value)
    return value


def _write_excel(result, handle) -> None:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", category=DeprecationWarning, module=r"openpyxl\..*"
        )
        from openpyxl import Workbook

        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Data"
        table = roi_table(result)
        data_sheet.append(table.columns)
        for row in table.rows:
            data_sheet.append(row)

        parameter_sheet = workbook.create_sheet("Parameters")
        parameter_sheet.append(("Parameter", "Value"))
        for name, value in result.parameters:
            parameter_sheet.append((name, _excel_value(value)))

        provenance_sheet = workbook.create_sheet("Provenance")
        provenance_sheet.append(("Field", "Value"))
        for name, value in vars(result.provenance).items():
            provenance_sheet.append((name, _excel_value(value)))

        warning_sheet = workbook.create_sheet("Warnings")
        warning_sheet.append(("Code", "Message", "Detail", "Field", "Remediation"))
        for warning in result.warnings:
            warning_sheet.append(
                (
                    warning.code.value,
                    warning.message,
                    warning.detail,
                    warning.field,
                    warning.remediation,
                )
            )
        workbook.save(handle)


def _failed_record(
    path: Path,
    suggested: str,
    format_name: str,
    code: str,
    message: str,
    detail: str,
) -> OutputFileRecord:
    issue = OutputIssue(
        severity=IssueSeverity.ERROR,
        code=code,
        message=message,
        detail=detail,
        path=str(path),
        format=format_name,
    )
    return OutputFileRecord(
        kind="data",
        format=format_name,
        path=str(path),
        suggested_filename=suggested,
        status=OutputFileStatus.FAILED,
        issues=(issue,),
    )


def _not_attempted_record(
    path: Path, suggested: str, format_name: str
) -> OutputFileRecord:
    issue = OutputIssue(
        severity=IssueSeverity.WARNING,
        code="not_attempted_after_failure",
        message="The file was not attempted.",
        detail="Failure policy stopped the request after an earlier output failed.",
        path=str(path),
        format=format_name,
    )
    return OutputFileRecord(
        kind="data",
        format=format_name,
        path=str(path),
        suggested_filename=suggested,
        status=OutputFileStatus.NOT_ATTEMPTED,
        issues=(issue,),
    )


def _attempt_write(
    request: ExportRequest,
    destination: Path,
    format: ExportFormat,
    stem: str,
) -> OutputFileRecord:
    suggested = f"{stem}_roi_data.{format.value}"
    target = destination / suggested
    existed = target.exists()
    if existed and request.overwrite is OverwritePolicy.FAIL_IF_EXISTS:
        return _failed_record(
            target,
            suggested,
            format.value,
            "target_exists",
            "The export target already exists.",
            "No file was changed because overwrite permission was not granted.",
        )

    temporary: Path | None = None
    created_target = False
    try:
        if request.overwrite is OverwritePolicy.REPLACE:
            descriptor, temp_name = tempfile.mkstemp(
                dir=destination, prefix=f".{suggested}.", suffix=".tmp"
            )
            os.close(descriptor)
            temporary = Path(temp_name)
            target_for_write = temporary
        else:
            target_for_write = target

        if format is ExportFormat.CSV:
            mode = "w" if temporary is not None else "x"
            handle = target_for_write.open(mode, newline="", encoding="utf-8")
            created_target = temporary is None
            with handle:
                _write_csv(request.result, handle)
        elif format is ExportFormat.EXCEL:
            mode = "wb" if temporary is not None else "xb"
            handle = target_for_write.open(mode)
            created_target = temporary is None
            with handle:
                _write_excel(request.result, handle)
        else:
            raise ValueError(f"Unsupported export format: {format!r}")

        if temporary is not None:
            os.replace(temporary, target)
    except Exception as exc:
        cleanup = temporary if temporary is not None else target
        if cleanup.exists() and (temporary is not None or created_target):
            cleanup.unlink()
        code = (
            "dependency_unavailable"
            if isinstance(exc, (ImportError, ModuleNotFoundError))
            else "target_exists"
            if isinstance(exc, FileExistsError)
            else "export_write_failed"
        )
        message = (
            "The export target already exists."
            if code == "target_exists"
            else "The export file could not be created."
        )
        return _failed_record(
            target,
            suggested,
            format.value,
            code,
            message,
            f"{type(exc).__name__}: {exc}",
        )

    issues = ()
    status = OutputFileStatus.CREATED
    if existed:
        status = OutputFileStatus.REPLACED
        issues = (
            OutputIssue(
                severity=IssueSeverity.WARNING,
                code="target_replaced",
                message="An existing export file was replaced.",
                detail="Replacement occurred under the explicit overwrite policy.",
                path=str(target),
                format=format.value,
            ),
        )
    return success_record(
        kind="data",
        format_name=format.value,
        path=target,
        suggested_filename=suggested,
        status=status,
        issues=issues,
    )


def export_roi_data(request: ExportRequest) -> OutputManifest:
    """Export only the requested formats and return an in-memory manifest."""
    if not request.formats:
        raise InvalidOutputRequestError(
            "no_export_formats",
            "At least one export format is required.",
            "formats",
        )
    if not all(isinstance(format, ExportFormat) for format in request.formats):
        raise InvalidOutputRequestError(
            "unsupported_export_format",
            f"Unsupported export formats: {request.formats!r}.",
            "formats",
        )
    if len(set(request.formats)) != len(request.formats):
        raise InvalidOutputRequestError(
            "duplicate_export_format",
            f"Each format may be requested once: {request.formats!r}.",
            "formats",
        )
    if not isinstance(request.overwrite, OverwritePolicy):
        raise InvalidOutputRequestError(
            "invalid_overwrite_policy",
            f"Unsupported overwrite policy: {request.overwrite!r}.",
            "overwrite",
        )
    if not isinstance(request.failure_policy, FailurePolicy):
        raise InvalidOutputRequestError(
            "invalid_failure_policy",
            f"Unsupported failure policy: {request.failure_policy!r}.",
            "failure_policy",
        )

    destination = prepare_destination(
        request.destination_directory, request.create_destination
    )
    stem = sanitize_stem(
        request.stem or request.result.suggested_export_stem,
        request.result.workflow_id,
    )
    records: list[OutputFileRecord] = []
    stopped = False
    for format in request.formats:
        suggested = f"{stem}_roi_data.{format.value}"
        target = destination / suggested
        if stopped:
            records.append(
                _not_attempted_record(target, suggested, format.value)
            )
            continue
        record = _attempt_write(request, destination, format, stem)
        records.append(record)
        if (
            record.status is OutputFileStatus.FAILED
            and request.failure_policy is FailurePolicy.STOP
        ):
            stopped = True
    return build_manifest(request.result.workflow_id, destination, records)
